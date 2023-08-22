""" Module for to convert differe file formats to Spark data frames. """

# core
import pandas as pd_legacy
from collections import Counter
import sys  # don't remove required for error handling
import os
import pyreadstat as prs

from urllib.parse import urlparse
from io import BytesIO

# types
import json
from collections import defaultdict

# libraries
from importlib import util

# adls and azure security
from azure.storage.filedatalake import DataLakeServiceClient
from azure.identity import ClientSecretCredential

# file excel
import openpyxl
import openpyxl.utils.cell

# PADE
import data_ecosystem_services.cdc_tech_environment_service.environment_file as cdc_env_file
import data_ecosystem_services.cdc_tech_environment_service.dataset_core as pade_ds_core
import data_ecosystem_services.cdc_security_service.security_core as pade_sec_core
import data_ecosystem_services.cdc_tech_environment_service.environment_core as pade_env_core


# data
import numpy as np

# spark / data
import uuid
from pyspark.sql import SparkSession, DataFrame
import pyspark.sql.utils
import pyspark.sql.functions as f
from pyspark.sql.functions import (
    col,
    lit,
    concat_ws,
    to_date,
    coalesce,
    trim,
    lpad,
    udf,
    length,
    when,
    expr,
    to_json,
    explode,
)

from pyspark.sql.types import StructType, IntegerType, StructField, LongType, ArrayType, StringType, DoubleType, DecimalType
from pyspark.sql import Row
uuid_udf = udf(lambda: str(uuid.uuid4()), StringType())

pyspark_pandas_loader = util.find_spec("pyspark.pandas")
pyspark_pandas_found = pyspark_pandas_loader is not None

if pyspark_pandas_found:
    os.environ['PYARROW_IGNORE_TIMEZONE'] = '1'
    import pyspark.pandas as pd
    # bug - pyspark version will not read local files
    # import pandas as pd
else:
    import pandas as pd


class DataSetConvert:
    """
    DataSetConvert Class to convert a dataset from one format to another.
    - Intended primarily to convert source file formats into Spark DataFrames.
    """

    @staticmethod
    def get_json_by_key(obj_json, target: str) -> dict:
        """Get json section based on the supplied path/target string

        Args:
            obj_json (any): Json object to search and extract section from
            target (str): Target/path to search for in json object

        Returns:
            any: Returns json section based on the supplied path/target string
        """

        for key in obj_json.keys():
            print(str(key))

        if target in obj_json:
            return obj_json[target]

        for key in obj_json.keys():
            key_scrubbed = key.strip().lower()
            target_scrubbed = target.strip().lower()
            print(
                f"looking for target_scrubbed:{target_scrubbed} in key_scrubbed:{key_scrubbed} "
            )
            if key_scrubbed == target_scrubbed:
                return obj_json[key]
        return {"message": "not_found", "key": target}

    @classmethod
    def convert_csv_tsv_usv_to_dataframe(cls, spark: SparkSession,
                                         ingress_file_path: str,
                                         first_row_is_header: bool,
                                         delimiter: str,
                                         encoding: str) -> DataFrame:
        """Loads csv or usv file to dataframe
        - Infers schema from file, postal codes will not be converted
        to strings
        - Date data types are not correctly inferred, TimeStamps are inferred
        - Column names are scrubbed to remove special characters and spaces
        - Original column names are stored in the metadata column

        Args:
            spark (SparkSession): SparkSession
            ingress_file_path (str): ADLS2 path to csv or usv file
            first_row_is_header (bool): true of false indicatory if file has
            header row
            delimiter (str): column delimiter
            encoding (str): encoding type

        Returns:
            DataFrame: Converted Spark dataframe
        """

        print("attempting to load dataframe for csv/tsv/usv")
        print(f"ingress_file_path:{ingress_file_path}")

        unsorted_df = (
            spark.read.format("csv")
            .option("header", first_row_is_header)
            .option("sep", delimiter)
            .option("encoding", encoding)
            .option("inferSchema", True)
            .option("treatEmptyValuesAsNulls", "true")
            .option("forceLowercaseNames", True)
            .option("multiline", "true")
            .option("inferLong", True)
            .option("inferDecimal", True)
            .option("inferInteger", True)
            .load(
                ingress_file_path,
                forceLowercaseNames=True,
                inferLong=True,
                inferDecimal=True,
                inferInteger=True,
            )
        )

        obj_ds_core = pade_ds_core.DataSetCore

        for col_orig in unsorted_df.columns:
            c_renamed = obj_ds_core.scrub_object_name(col_orig)
            unsorted_df = unsorted_df.withColumnRenamed(col_orig, c_renamed)
            unsorted_df = unsorted_df.withMetadata(
                c_renamed, {"ingress_column_name": col_orig, 'comment': col_orig})
            print(
                f"{c_renamed} :metadata: {unsorted_df.schema[c_renamed].metadata}")

        return unsorted_df

    @classmethod
    def convert_json_to_dataframe(cls, spark: SparkSession, ingress_file_path: str, encoding: str,
                                  source_json_path: str, source_dataset_name: str, dataset_name: str) -> DataFrame:
        """Loads json file to dataframe
        - Infers schema from file, postal codes will not be converted to strings
        - Date data types are not correctly inferred, TimeStamps are inferred
        - Column names are created based on the supplied configuration
        - Original column names are stored in the metadata column
        - The system currently supports nested and flat json.
        - The structure of the json is inferred based on the number of rows/columns found at the specified path

        Args:
            spark (SparkSession): SparkSession
            ingress_file_path (str): ADLS2 path to json file
            encoding (str): Encoding type
            source_json_path (str): Json path to section of json file to be loaded
            source_dataset_name (str): Name of the source dataset - name in json
            dataset_name (str): Name of the dataset/table to create

        Returns:
            DataFrame: Converted Spark dataframe
        """

        print("attempting to load dataframe for json")
        print(f"ingress_file_path:{ingress_file_path}")
        json_df = (
            spark.read.format("json")
            .option("multiline", "true")
            .option("encoding", encoding)
            .option("inferSchema", True)
            .option("forceLowercaseNames", True)
            .option("inferLong", True)
            .option("inferDecimal", True)
            .option("inferInteger", True)
            .load(
                ingress_file_path,
                forceLowercaseNames=True,
                inferLong=True,
                inferDecimal=True,
                inferInteger=True,
            )
        )

        # get column with the list of json for that type
        if len(source_json_path.strip()) > 0:
            # nested json
            list_columns = json_df.columns
            json_df = json_df.withColumn(
                "json_column", to_json(list_columns[0]))
            json_df_to_pandas = json_df.toPandas()
            first_row = json_df_to_pandas["json_column"].iloc[0]
            json_data = json.loads(first_row)
            dict_json = cls.get_json_by_key(json_data, source_dataset_name)

            dict_list = dict_json.items()
            dict_dataframe = defaultdict(list)
            num_cols = 0
            for i, col_list in enumerate(dict_list):
                num_cols = len(col_list) + 1
                values = list(col_list)
                values.append(source_dataset_name)
                tuble_values = tuple(values)
                dict_dataframe[tuble_values].append(i)

            list_dataframe = list(dict_dataframe)

            new_cols = []
            for c_orig in range(1, num_cols):
                c_renamed = "column" + str(c_orig + 1)
                new_cols.append(c_renamed)
            new_cols.append("source_dataset_name")

            print(f"num_cols:{num_cols} new_cols:{new_cols}")
            print(f" dict_list:{str(dict_list)}")

            df_json = pd_legacy.DataFrame(list_dataframe, columns=new_cols)
            df_json.columns = new_cols
            print(df_json)
            unsorted_df = spark.createDataFrame(df_json)
            unsorted_df.na.fill(0)
            unsorted_df.na.fill("")
            unsorted_df.show()

        else:
            # flat json
            print(f"ingress_file_path:{ingress_file_path}")
            json_df_to_pandas = json_df.toPandas()
            json_string = json_df_to_pandas.to_json(orient="index")

            json_data = json.loads(json_string)
            dict_json = cls.get_json_by_key(json_data, "0")
            source_dataset_name = dataset_name
            dict_list = dict_json.items()
            dict_dataframe = defaultdict(list)
            num_cols = 0
            for i, x in enumerate(dict_list):
                num_cols = len(x) + 1
                values = list(x)
                values.append(source_dataset_name)
                tuble_values = tuple(values)
                dict_dataframe[tuble_values].append(i)

            list_dataframe = list(dict_dataframe)

            new_cols = []

            for c_orig in range(1, num_cols):
                c_renamed = "column" + str(c_orig + 1)
                new_cols.append(c_renamed)

            new_cols.append("source_dataset_name")

            print(f"num_cols:{num_cols} new_cols:{new_cols}")
            print(f" dict_list:{str(dict_list)}")

            df_json = pd_legacy.DataFrame(list_dataframe, columns=new_cols)
            df_json.columns = new_cols
            print(df_json)
            unsorted_df = spark.createDataFrame(df_json)
            unsorted_df.na.fill(0)
            unsorted_df.na.fill("")

        return unsorted_df

    # ingress_mount = f"/mnt/ddnid-nccdphp/OD/NHanes/dev/ingress"
    @classmethod
    def convert_sas_to_dataframe_with_schema(cls, spark: SparkSession, ingress_file_path: str, tenant_id: str,
                                             client_id: str, client_secret: str):
        """ Loads xpt or bdat to dataframe and extracts schema from xpt file
        - Reference article: https://stackoverflow.com/questions/58612304/reading-huge-sas-dataset-in-python
        - PS: Please be noted that the resulting dataframe allChunk is going to have all column as Categorical data

        Args:
            spark (SparkSession): SparkSession
            ingress_file_path (str): ADLS2 path to csv or usv file
            tenant_id (str): azure tenant id used to download file
            client_id (str): azure client id used to download file
            client_secret (str): azure secret used to download file

        Returns:
            DataFrame: Converted Spark dataframe, schema is extracted from xpt file
        """

        obj_env_file = cdc_env_file.EnvironmentFile()
        https_path = obj_env_file.convert_abfss_to_https_path(
            ingress_file_path)
        credential = ClientSecretCredential(
            tenant_id, client_id, client_secret)

        # download file
        storage_account_loc = urlparse(https_path).netloc
        account_url = f"https://{storage_account_loc}"
        storage_path = urlparse(https_path).path
        storage_container = storage_path.split('/')[1]
        file_path = storage_path.replace(f"{storage_container}" + "/", "")
        print(f"storage_path:{storage_path}")
        print(f"https_path:{https_path}")
        # print(f"ingress_mount:{ingress_mount}")
        print(f"ingress_file_path:{ingress_file_path}")
        service_client = DataLakeServiceClient(
            account_url=account_url, credential=credential)
        file_system_client = service_client.get_file_system_client(
            storage_container)
        file_client = file_system_client.get_file_client(file_path)
        download = file_client.download_file(0)
        download_bytes = download.readall()
        download_file = BytesIO(download_bytes)

        # step 1: get pandas data frame

        chunk_size = 50000
        offset = 0
        # Get the function object in a variable getChunk
        if file_path.lower().endswith('sas7bdat'):
            get_chunk = prs.read_sas7bdat
        else:
            get_chunk = prs.read_xport

        df_all_chunks, _ = get_chunk(
            ingress_file_path, row_limit=chunk_size, row_offset=offset)
        df_all_chunks = df_all_chunks.astype('category')

        while True:
            offset += chunk_size
            # for xpt data, use pyreadstat.read_xpt()
            chunk, _ = prs.read_xport(
                ingress_file_path, row_limit=chunk_size, row_offset=offset)
            # if chunk is empty, it means the entire data has been read, so break
            if chunk.empty:
                break

            # converting each column to categorical
            for each_col in chunk:
                col_union = pd_legacy.api.types.union_categoricals(
                    [df_all_chunks[each_col], chunk[each_col]])
                df_all_chunks[each_col] = pd_legacy.Categorical(df_all_chunks[each_col],
                                                                categories=col_union.categories)
                chunk[each_col] = pd_legacy.Categorical(
                    chunk[each_col], categories=col_union.categories)

            # Append each chunk to the resulting dataframe
            df_all_chunks = pd_legacy.concat([df_all_chunks, chunk])

        df_xpt, xpt_meta = prs.read_xport(download_file)
        df_data = spark.createDataFrame(df_xpt)

        # step 2: initalize empty pandas datafram
        df_metadata = pd.DataFrame()

        # read column name, labels into the new pandas dataframe
        df_metadata["name"] = xpt_meta.column_names = xpt_meta.column_names.str.strip()
        df_metadata["label"] = xpt_meta.column_labels = [
            x.decode('utf-8') for x in xpt_meta.column_labels]
        df_metadata["format"] = xpt_meta.column_formats = [
            x.decode('utf-8') for x in xpt_meta.column_formats]
        df_metadata["type"] = xpt_meta.column_types = [
            x.decode('utf-8') for x in xpt_meta.variable_types]
        df_metadata["type_in_source"] = xpt_meta.column_types = [
            x.decode('utf-8') for x in xpt_meta.original_variable_types]
        df_metadata["length"] = xpt_meta.column_lengths = xpt_meta.column_lengths.astype(
            int)
        df_metadata["note"] = xpt_meta.column_notes = [
            x.decode('utf-8') for x in xpt_meta.column_notes]
        df_metadata["table_name"] = xpt_meta.table_name = xpt_meta.table_name.decode(
            'utf-8')

        df_metadata = spark.createDataFrame(df_metadata)
        df_metadata.show()
        # extracting number of rows from the Dataframe
        row_count = df_metadata.count()
        # extracting number of columns from the Dataframe
        col_count = len(df_metadata.columns)
        # printing
        print(f"Dimension of the Dataframe is: {(row_count, col_count)}")
        print(f"Number of Rows are: {row_count}")
        print(f"Number of Columns are: {col_count}")
        return df_data, df_metadata

    @classmethod
    def pandas_to_spark(cls, df_pandas, spark):
        """_summary_: Converts pandas dataframe to spark dataframe

        Args:
            df_pandas (_type_): _description_
            spark (_type_): _description_

        Returns:
            _type_: _description_
        """

        list_pandas = []
        df_pandas = df_pandas.reset_index()  # make sure indexes pair with number of rows
        for index, row in df_pandas.iterrows():
            print(row['column_name'], row['comment'])
            row_to_append = (row['column_name'], row['comment'])
            list_pandas.append(row_to_append)
        rdd = spark.sparkContext.parallelize(list_pandas)
        rdd_columns = ["column_name", "comment"]
        df_spark = rdd.toDF(rdd_columns)
        return df_spark

    @classmethod
    def convert_sas_mount_to_dataframe_with_schema(cls, spark: SparkSession, source_abbreviation: str, file_name: str, tenant_id: str,
                                                   client_id: str, client_secret: str):
        """ Loads xpt or bdat to dataframe and extracts schema from xpt file
        - Reference article: https://stackoverflow.com/questions/58612304/reading-huge-sas-dataset-in-python
        - PS: Please be noted that the resulting dataframe allChunk is going to have all column as Categorical data

        Args:
            spark (SparkSession): SparkSession
            ingress_file_path (str): ADLS2 path to csv or usv file
            tenant_id (str): azure tenant id used to download file
            client_id (str): azure client id used to download file
            client_secret (str): azure secret used to download file

        Returns:
            DataFrame: Converted Spark dataframe, schema is extracted from xpt file
        """

        # obj_ds_core = pade_ds_core.DataSetCore()
        obj_env_file = cdc_env_file.EnvironmentFile()
        mount_file_path = "/Workspace/Repos/tbi8@cdc.gov/pade/od/od_nhanes/temp/"
        ingress_path = f"{mount_file_path}{source_abbreviation}/{file_name}"
        print(f"mount_file_path:{mount_file_path}")
        print(f"ingress_path:{ingress_path}")

        # df_xpt, xpt_meta = prs.read_xport(ingress_path,metadataonly=True)
        df_xpt, xpt_meta = prs.read_xport(ingress_path)

        # column_name                    xpt_meta.column_names
        # data_type_name                 xpt_meta.variable_types
        # comment                        xpt_meta.column_notes.decode('utf-8')
        # dataset_name
        # full_dataset_name
        # dataset_file_path
        # row_id
        # unique_count
        # max_length
        # min_length
        # ingress_column_name            xpt_meta.column_names
        # ingress_column_format          xpt_meta.column_formats
        # ingress_column_label           xpt_meta.column_labels
        # unique_count_scrubbed
        # scope
        # row_id_column
        # row_count                      xpt_meta.number_rows.astype(int)
        # ingress_row_count              xpt_meta.number_rows.astype(int)
        # ingress_ordinal_position
        # ingress_column_length           xpt_meta.column_lengths.astype(int)
        # ingress_table_name              xpt_meta.table_name.decode('utf-8')

        list_pandas = list(zip(xpt_meta.column_names, xpt_meta.column_labels))
        print(f"list_pandas: {list_pandas}")
        rdd = spark.sparkContext.parallelize(list_pandas)

        schema = StructType(
            [
                StructField("column_name", StringType(), False),
                StructField("data_type_name", StringType(), False),
                StructField("comment", StringType(), True),
                StructField("dataset_name", StringType(), False),
                StructField("full_dataset_name", StringType(), False),
                StructField("dataset_file_path", StringType(), False),
                StructField("row_id", StringType(), False),
                StructField("unique_count", LongType(), True),
                StructField("null_count", LongType(), True),
                StructField("max_length", LongType(), True),
                StructField("min_length", LongType(), True),
                StructField("ingress_column_name", StringType(), True),
                StructField("ingress_column_format", StringType(), True),
                StructField("ingress_column_label", StringType(), True),
                StructField("unique_count_scrubbed", LongType(), True),
                StructField("scope", StringType(), True),
                StructField("row_id_column", StringType(), True),
            ]
        )

        schema_rdd = StructType([StructField("column_name", StringType(), False),
                                 StructField("column_label", StringType(), True)])
        df_metadata_spark = spark.createDataFrame(rdd, schema=schema_rdd)
        df_metadata_spark.printSchema()
        df_metadata_spark.show(truncate=False)

        # extracting number of rows from the Dataframe
        row = df_metadata_spark.count()
        # extracting number of columns from the Dataframe
        col = len(df_metadata_spark.columns)
        # printing
        print(f"Dimension of the Dataframe is: {(row,col)}")
        print(f"Number of Rows are: {row}")
        print(f"Number of Columns are: {col}")

        return df_metadata_spark, df_xpt

    @classmethod
    def convert_xpt_to_dataframe(cls, spark: SparkSession, ingress_file_path: str, tenant_id: str, client_id: str,
                                 client_secret: str) -> DataFrame:
        """Loads xpt to dataframe
        - XPT supports types - XPT types are converted to Spark types
        - Column names are not scrubbed or changed
        - No metadata is added to the dataframe
        - Pandas does not support accessing abfss directly so file is downloaded prior to loading

        Args:
            spark (SparkSession): SparkSession
            ingress_file_path (str): ADLS2 path to csv or usv file
            tenant_id (str): azure tenant id used to download file
            client_id (str): azure client id used to download file
            client_secret (str): azure secret used to download file

        Returns:
            DataFrame: Spark dataframe
        """

        obj_env_file = cdc_env_file.EnvironmentFile()
        https_path = obj_env_file.convert_abfss_to_https_path(
            ingress_file_path)
        credential = ClientSecretCredential(
            tenant_id, client_id, client_secret)

        # download file
        storage_account_loc = urlparse(https_path).netloc
        account_url = f"https://{storage_account_loc}"
        storage_path = urlparse(https_path).path
        storage_container = storage_path.split('/')[1]
        file_path = storage_path.replace(f"{storage_container}" + "/", "")
        service_client = DataLakeServiceClient(
            account_url=account_url, credential=credential)
        file_system_client = service_client.get_file_system_client(
            storage_container)
        file_client = file_system_client.get_file_client(file_path)
        download = file_client.download_file(0)
        download_bytes = download.readall()
        download_file = BytesIO(download_bytes)

        # step 1: get pandas data frame
        df_xpt = pd_legacy.read_sas(download_file, format='xport')
        # step 2: get spark data frame
        unsorted_df = spark.createDataframe(df_xpt)

        return unsorted_df

    @classmethod
    def convert_sas_to_dataframe(cls, spark: SparkSession, ingress_file_path: str, tenant_id: str, client_id: str,
                                 client_secret: str) -> DataFrame:
        """Loads sas dta to dataframe
        - XPT supports types - XPT types are converted to Spark types
        - Column names are not scrubbed or changed
        - No metadata is added to the dataframe
        - Pandas does not support accessing abfss directly so file is downloaded prior to loading

        Args:
            spark (SparkSession): SparkSession
            ingress_file_path (str): ADLS2 path to csv or usv file
            tenant_id (str): azure tenant id used to download file
            client_id (str): azure client id used to download file
            client_secret (str): azure secret used to download file

        Returns:
            DataFrame: Spark dataframe
        """

        obj_env_file = cdc_env_file.EnvironmentFile()
        https_path = obj_env_file.convert_abfss_to_https_path(
            ingress_file_path)
        credential = ClientSecretCredential(
            tenant_id, client_id, client_secret)

        # download file
        storage_account_loc = urlparse(https_path).netloc
        account_url = f"https://{storage_account_loc}"
        storage_path = urlparse(https_path).path
        storage_container = storage_path.split('/')[1]
        file_path = storage_path.replace(f"{storage_container}" + "/", "")
        service_client = DataLakeServiceClient(
            account_url=account_url, credential=credential)
        file_system_client = service_client.get_file_system_client(
            storage_container)
        file_client = file_system_client.get_file_client(file_path)
        download = file_client.download_file(0)
        download_bytes = download.readall()
        download_file = BytesIO(download_bytes)

        # step 1: get pandas data frame
        df_sas = pd_legacy.read_sas(download_file, format='sas7bdat')
        # step 2: get spark data frame
        unsorted_df = spark.createDataFrame(df_sas)
        return unsorted_df

    @classmethod
    def convert_xlsx_to_dataframe(cls, spark: SparkSession, ingress_file_path: str, sheet_name: str, tenant_id: str,
                                  client_id: str, client_secret: str,
                                  skip_rows_text: str) -> DataFrame:
        """Loads xlsx file to dataframe
        - Currently supports xlsx only not xls
        - Supports both formulas and values
        - Supports url extraction
        - All data is imported as strings
        - Original column names are stored in the metadata column
        - Column names are scrubbed for special characters and spaces

        Args:
            spark (SparkSession): SparkSession
            ingress_file_path (str): ADLS2 path to xlsx file
            sheet_name (str): name of the worksheet or ordinal position starting at 0
            tenant_id (str): azure tenant id used to download file
            client_id (str): azure client id used to download file
            client_secret (str): azure secret used to download file
            skip_rows (str): number of rows to skip at the top of the spreadsheet before headers

        Returns:
            DataFrame: Converted Spark dataframe
        """

        obj_ds_core = pade_ds_core.DataSetCore()
        print("attempting to load dataframe for xlsx")
        print(f"ingress_file_path:{ingress_file_path}")

        if skip_rows_text is None:
            skip_rows = 0
        else:
            if str(skip_rows_text) == "":
                skip_rows = 0
            elif len(str(skip_rows_text).strip()) == 0:
                skip_rows = 0
            else:
                skip_rows = int(skip_rows_text)

        if sheet_name is None:
            sheet_name = ""

        obj_env_file = cdc_env_file.EnvironmentFile()
        https_path = obj_env_file.convert_abfss_to_https_path(
            ingress_file_path)
        credential = ClientSecretCredential(
            tenant_id, client_id, client_secret)

        # download file
        storage_account_loc = urlparse(https_path).netloc
        account_url = f"https://{storage_account_loc}"
        storage_path = urlparse(https_path).path
        storage_container = storage_path.split('/')[1]
        file_path = storage_path.replace(f"{storage_container}" + "/", "")
        service_client = DataLakeServiceClient(
            account_url=account_url, credential=credential)
        file_system_client = service_client.get_file_system_client(
            storage_container)
        file_client = file_system_client.get_file_client(file_path)
        download = file_client.download_file(0)
        download_bytes = download.readall()
        download_file = BytesIO(download_bytes)

        work_book_data = None
        print("try open file")
        print("try load workbook")
        print(f"file_path:{file_path}")

        if sheet_name is None or sheet_name == "":
            print("try read_excel")
            df_excel = pd_legacy.read_excel(
                download_file,
                skiprows=skip_rows,
                engine="openpyxl",
                dtype="string",
                na_filter=False,
                header=None,
            )
        else:
            sheet_name = str(sheet_name)
            print("try read_excel")
            if sheet_name.isdigit():
                df_excel = pd_legacy.read_excel(
                    download_file,
                    skiprows=skip_rows,
                    engine="openpyxl",
                    dtype="string",
                    sheet_name=int(sheet_name),
                    na_filter=False,
                    header=None,
                )
            else:
                df_excel = pd_legacy.read_excel(
                    download_file,
                    skiprows=skip_rows,
                    engine="openpyxl",
                    dtype="string",
                    sheet_name=sheet_name,
                    na_filter=False,
                    header=None,
                )

        data = df_excel.values.tolist()
        cols = np.asarray(data[0:1][0])
        data = data[skip_rows + 1:]
        print(f"original cols:{cols}")
        cols = cols.tolist()
        cols = obj_ds_core.rename_column_names_as_unique(cols)
        print(f"deduplicated cols:{cols}")
        print(f"len(cols):{len(cols)}")
        print("call loadworkbook data")
        print(
            f"2 open {https_path} work_book_data data_only False: read_only True")

        try:
            work_book_data = openpyxl.load_workbook(
                download_file, data_only=False, read_only=False
            )
            g_sheet = work_book_data.sheetnames
            if sheet_name is None or sheet_name == "":
                work_sheet_data = work_book_data[0]
            else:
                if sheet_name.isdigit():
                    work_sheet_data = work_book_data[g_sheet[int(sheet_name)]]
                else:
                    work_sheet_data = work_book_data[sheet_name]

            data_pass_1 = cls.convert_xlsx_to_list(
                data, cols, work_sheet_data, skip_rows)
            print("close work_book_data data_only True: read_only True")
        finally:
            if work_book_data is not None:
                work_book_data.close()

        # formula
        print(f"3 open {https_path} work_book_formula")
        print("data_only False: read_only False")
        work_book_formula = None
        try:
            work_book_formula = openpyxl.load_workbook(
                download_file, data_only=True, read_only=False
            )
            g_sheet = work_book_formula.sheetnames
            if sheet_name is None or sheet_name == "":
                work_sheet_formula = work_book_formula[0]
            else:
                if sheet_name.isdigit():
                    work_sheet_formula = work_book_formula[g_sheet[int(
                        sheet_name)]]
                else:
                    work_sheet_formula = work_book_formula[sheet_name]

            print("call loadworkbook formula")
            data_pass_2 = cls.convert_xlsx_to_list(
                data_pass_1, cols, work_sheet_formula, skip_rows
            )

        finally:
            if work_book_formula is not None:
                work_book_formula.close()

        print("close work_book_formula data_only False: read_only False")
        # data = ( islice(r, i , None) for r in data)

        obj_ds_core = pade_ds_core.DataSetCore

        df_excel = pd_legacy.DataFrame(data_pass_2, columns=cols)
        df_excel = df_excel.astype(str)
        # df_excel.columns = df_excel.columns.str.replace(" ", "_")
        df_excel = df_excel.replace("nan", "")
        df_excel.replace(np.NaN, "", inplace=True)
        print(df_excel)
        unsorted_df = spark.createDataFrame(df_excel)
        for col_orig in unsorted_df.columns:
            # Data to replace format: Unnamed:_1
            c_renamed = obj_ds_core.scrub_object_name(col_orig)
            if c_renamed.startswith("unnamed:_"):
                c_renamed = c_renamed.replace("unnamed:_", "")
                # increments #by 1
                c_renamed = int(c_renamed) + 1
                c_renamed = "Column" + str(c_renamed)
            unsorted_df = unsorted_df.withColumnRenamed(col_orig, c_renamed)
            unsorted_df = unsorted_df.withMetadata(
                c_renamed, {"ingress_column_name": col_orig, 'comment': col_orig})

        # replace nulls
        unsorted_df.na.fill(0)
        unsorted_df.na.fill("")
        unsorted_df.show()
        unsorted_df = unsorted_df.withColumn(
            "__meta_sheet_name", lit(sheet_name))

        return unsorted_df

    @staticmethod
    def convert_xlsx_to_list(data, cols, work_sheet, skip_rows):
        """Loads excel worksheet into a list of dictionaries

        Args:
            data (_type_): _description_
            cols (_type_): _description_
            work_sheet (_type_): _description_
            skip_rows (_type_): _description_

        Returns:
            list: Python list with converted Excel Data
        """

        # iterate through excel and display data
        for i_row, row_item in enumerate(data):
            row_item_updated = [""] * (len(cols))
            for j_col, col_item in enumerate(row_item):
                col_item_updated = col_item
                cell = work_sheet.cell(
                    row=i_row + 2 + skip_rows, column=j_col + 1)
                if cell.hyperlink is not None:
                    url = str(cell.hyperlink.target)
                    col_item_updated = "[" + str(cell.value) + "](" + url + ")"
                else:
                    if cell.value is not None:
                        col_item_updated = cell.value
                    else:
                        col_item_updated = ""
                row_item_updated[j_col] = col_item_updated
                print(f"col_item_updated:{col_item_updated}")
            data[i_row] = row_item_updated
        return data
