#! /usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Python Version: 3.11.4
Creator: 孙一凡-莫无煜
Create Date: 2023/7/24

"""
import json

import openpyxl

a = openpyxl.Workbook()
# b = a.active
# a.remove(b)
s = a.active
with open('C:\\Users\\movoid\\Documents\\WeChat Files\\wxid_zrhiq7za1bhf22\\FileStorage\\File\\2023-07\\log_analysed.json', mode='r') as file:
    all_dict = json.load(file)['02']
print(all_dict)
index = 0
for i, v in all_dict.items():
    # s = a.create_sheet(i)

    s.cell(index * 3 + 1, 1, 'Micro Site')
    s.cell(index * 3 + 2, 1, i)
    s.cell(index * 3 + 1, 2, 'total time')
    s.cell(index * 3 + 2, 2, sum([_[-1] for _ in v]))
    s.cell(index * 3 + 1, 3, 'count')
    s.cell(index * 3 + 2, 3, len(v))
    for j, w in enumerate(v):
        s.cell(index * 3 + 1, j + 4, w[0])
        s.cell(index * 3 + 2, j + 4, w[-1])
    index += 1

a.save('log_one.xlsx')

if __name__ == '__main__':
    pass
